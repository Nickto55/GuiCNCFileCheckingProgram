import os
from datetime import date
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from useJson import JsonConfig, JsonSave

listSpli = []
Recursion_Depth = 2
today = date.today()


class MainAutomaticallySearch:
    pass

def auto_fit_columns(sheet):
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        adjusted_width = (max_length - 20)  # Немного увеличим для красоты
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['E'].width = 1
    sheet.column_dimensions['F'].width = 5
    sheet.column_dimensions['H'].width = 1
    sheet.column_dimensions['K'].width = 1
    sheet.column_dimensions['J'].width = 5


def collect_subdirectories(root_dir_list: list, progress_tracker=None, lastTimeAuvtoSearchBool: bool = False):
    global Recursion_Depth, today
    result = {}

    # Устанавливаем общее количество шагов для прогресса (количество директорий)
    if progress_tracker:
        progress_tracker.set_total(len(root_dir_list))

    count = 0

    for dir_name in root_dir_list:
        # Обновляем прогресс в начале обработки директории
        # Передаем count (0-based index) и сообщение
        if progress_tracker:
            progress_tracker.update(count, f"Обработка директории: {os.path.basename(dir_name)}")

        project_path = dir_name
        print(f"Создание листа: {os.path.basename(dir_name)}...")

        subdirs = []

        def recurse(path):
            global Recursion_Depth
            try:
                for item in os.listdir(path):
                    full_path = os.path.join(path, item)
                    if os.path.isdir(full_path):
                        if full_path.count("\\") == Recursion_Depth:
                            subdirs.append({'name': item, 'content': "", 'full_path': full_path})
                        recurse(full_path)
                    if full_path.count("\\") == Recursion_Depth + 1:
                        parent_dir = os.path.basename(os.path.dirname(full_path))
                        subdirs.append({'name': parent_dir, 'content': item, 'full_path': full_path})

            except PermissionError:
                print(f"Не удалось получить доступ к {path}")

        recurse(project_path)

        result[os.path.basename(dir_name)] = subdirs
        count += 1

    # Завершаем прогресс (устанавливаем на 100% или последний шаг)
    if progress_tracker:
        progress_tracker.update(len(root_dir_list) - 1 if root_dir_list else 0, "Обработка директорий завершена")

    return result


def save_to_excel(data, output_file):
    column_indices = [1, 2, 3]
    data_range = "A1:D100"  # Увеличиваем диапазон для фильтра

    fill_color = PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")

    wb = openpyxl.Workbook()
    del wb["Sheet"]

    for project_name, dirs in data.items():
        # Ограничиваем длину имени листа до 31 символа (ограничение Excel)
        safe_sheet_name = project_name[:31]
        # Убираем недопустимые символы для имени листа Excel
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        for char in invalid_chars:
            safe_sheet_name = safe_sheet_name.replace(char, '_')

        ws = wb.create_sheet(title=safe_sheet_name)

        # Заголовки
        headers = ["", "Название ДСЕ", "Содержимое", "Путь", "", "Fm", "Файлы без расширения", "", "Дата последнего изменения", " KБ", ""]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # создание фильтра и его настройка
        # Корректируем диапазон для фильтра: от A1 до последнего столбца и достаточного количества строк
        if dirs:
            last_row = len(dirs) + 1  # +1 для заголовка
        else:
            last_row = 1
        filter_range = f"A1:{get_column_letter(len(headers))}{last_row}"
        ws.auto_filter.ref = filter_range

        dirDseList = [2, 0]  # [row, boolFm]
        # Заполнение данных
        for row_idx, entry in enumerate(dirs, 2):
            # Убедимся, что ключи существуют
            name = entry.get('name', '')
            content = entry.get('content', '')
            full_path = entry.get('full_path', '')

            ws.cell(row=row_idx, column=1, value="")  # Пустая колонка A
            ws.cell(row=row_idx, column=2, value=name)
            cell_content = ws.cell(row=row_idx, column=3, value=content)

            # Расскраска, если content пустой
            if not content:
                for col_num in range(1, len(headers) + 1):  # Используем len(headers)
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.fill = fill_color
                    try:
                        if dirDseList[1] == 1:
                            for rowDirFm in range(dirDseList[0], row_idx):
                                ws.cell(row=rowDirFm, column=6, value="X")

                        dirDseList = [row_idx, 0]
                    except:
                        dirDseList = [row_idx, 0]
                        print("Ошибка: dirDseList пуст")

            # Создание гиперссылки
            cell_path = ws.cell(row=row_idx, column=4, value=full_path)
            # Проверяем, что full_path не пустой перед созданием гиперссылки
            if full_path:
                cell_path.hyperlink = full_path

                file_name, file_extension = os.path.splitext(full_path)

                if content and file_extension == "" and not (os.path.isdir(full_path)):
                    ws.cell(row=row_idx, column=7, value="X")

                if file_extension == ".fm" and dirDseList[1] != 1:
                    dirDseList.pop()
                    dirDseList.append(1)

                # print(os.path.basename(full_path))
                # Можно установить стиль для гиперссылок, если нужно
                # from openpyxl.styles import Font
                # cell_path.font = Font(color="0000FF", underline="single")

        # Регулировка ширины столбцов
        auto_fit_columns(ws)
        # Закрепление первой строки
        ws.freeze_panes = 'A2'
        ws.column_dimensions['E'].fill = fill_color
        ws.column_dimensions['K'].fill = fill_color
        ws.column_dimensions['H'].fill = fill_color

    try:
        wb.save(output_file)
        print(f"Файл сохранен: {output_file}")
    except Exception as e:
        print(f"Ошибка при сохранении файла {output_file}: {e}")
        raise  # Перебрасываем исключение


def mainCNCFileCheckingProgram(list_main_repo: list, choseUser: int, twoProgramm: int, progress_tracker=None,
                               countProg=1):
    global listSpli
    global Recursion_Depth

    # Получаем текущую директорию, где находится скрипт
    current_directory = os.getcwd()
    CONFIG_DIR = os.path.join(os.path.expanduser("~"), ".CNCFileCheckingProgram")

    def nameUserSee(CONFIG_DIR):
        global nameUser
        nameUser = ""
        count = 0

        for i in CONFIG_DIR:
            if i == os.sep:  # Используем os.sep
                if count < 3:
                    count += 1
                else:
                    break
            if count == 2 and i != os.sep:
                nameUser += i

    nameUserSee(CONFIG_DIR)
    # --- УЛУЧШЕННАЯ ПРОВЕРКА ---
    if not list_main_repo:
        print(f"Не указано ни одного репозитория.")
        if twoProgramm:
            output_file = f"BD_CNCprog_{today}.xlsx"
            full_output_path = os.path.join(current_directory, output_file)
            try:
                wb = openpyxl.Workbook()
                del wb["Sheet"]  # Удаляем лист по умолчанию
                wb.save(full_output_path)
                print(f"Создан пустой файл: {full_output_path}")
                return output_file  # Возвращаем имя файла
            except Exception as e:
                print(f"Ошибка при создании пустого файла: {e}")
                return None  # Или поднять исключение
        else:
            return

    Recursion_Depth = 2

    # Передаем progress_tracker в collect_subdirectories
    data = collect_subdirectories(list_main_repo, progress_tracker)  # Убираем countProg отсюда

    output_file = f"BD_CNCprog_{today}"
    if not output_file.endswith(".xlsx"):
        output_file += ".xlsx"

    full_output_path = os.path.join(current_directory, output_file)

    # Оборачиваем save_to_excel в try-except
    try:
        save_to_excel(data, full_output_path)
    except Exception as e:
        print(f"Ошибка при сохранении файла в mainCNCFileCheckingProgram: {e}")
        # В зависимости от требований, можно поднять исключение или вернуть None
        if twoProgramm:
            return None  # Или поднять исключение, если вызывающая сторона должна его обработать
        else:
            return  # Просто завершаем

    return output_file  # Возвращаем имя файла для следующей программы

if __name__ == "__main__":
    collect_subdirectories([r"C:\Users\yakovlev_nd\Desktop\Test\CNCFileCheckingProgram TestDir\2Dahlih", r"C:\Users\yakovlev_nd\Desktop\Test\CNCFileCheckingProgram TestDir\3Dahlih"], None, True)

import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

class ExcelEnter:
    def __init__(self):
        pass

    @staticmethod
    def auto_fit_columns(sheet):
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            adjusted_width =max_length if (max_length - 20) < 30 else 20
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['D'].width = 80
        sheet.column_dimensions['E'].width = 1
        sheet.column_dimensions['F'].width = 5
        sheet.column_dimensions['H'].width = 1
        sheet.column_dimensions['J'].width = 5
        sheet.column_dimensions['K'].width = 1

    def save_to_excel(self, data, output_file):
        fill_color = PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")

        wb = openpyxl.Workbook()
        del wb["Sheet"]

        for project_name, dirs in data.items():
            # Ограничиваем длину имени листа до 31 символа (ограничение Excel)
            safe_sheet_name = project_name[:31]
            invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
            for char in invalid_chars:
                safe_sheet_name = safe_sheet_name.replace(char, '_')

            ws = wb.create_sheet(title=safe_sheet_name)

            # Заголовки
            headers = ["", "Название ДСЕ", "Содержимое", "Путь", "", "Fm", "Файлы без расширения", "", "Дата последнего изменения", "KБ", ""]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            # создание фильтра и его настройка
            if dirs:
                last_row = len(dirs) + 1
            else:
                last_row = 1
            filter_range = f"A1:{get_column_letter(len(headers))}{last_row}"
            ws.auto_filter.ref = filter_range

            dir_dse_list = [2, 0]
            for row_idx, entry in enumerate(dirs, 2):
                name = entry.get('name', '')
                content = entry.get('content', '')
                full_path = entry.get('full_path', '')

                ws.cell(row=row_idx, column=1, value="")
                ws.cell(row=row_idx, column=2, value=name)
                ws.cell(row=row_idx, column=3, value=content)

                # Расскраска, если content пустой
                if not content:
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=col_num)
                        cell.fill = fill_color
                        try:
                            if dir_dse_list[1] == 1:
                                for rowDirFm in range(dir_dse_list[0], row_idx):
                                    ws.cell(row=rowDirFm, column=6, value="X")

                            dir_dse_list = [row_idx, 0]
                        except:
                            dir_dse_list = [row_idx, 0]
                            print("Ошибка: dir_dse_list пуст")

                # Создание гиперссылки
                cell_path = ws.cell(row=row_idx, column=4, value=full_path)
                if full_path:
                    cell_path.hyperlink = full_path

                    file_name, file_extension = os.path.splitext(full_path)

                    if content and file_extension == "" and not (os.path.isdir(full_path)):
                        ws.cell(row=row_idx, column=7, value="X")

                    if file_extension == ".fm" and dir_dse_list[1] != 1:
                        dir_dse_list.pop()
                        dir_dse_list.append(1)

            # Регулировка ширины столбцов
            self.auto_fit_columns(ws)
            # Закрепление первой строки
            ws.freeze_panes = 'A2'

            ws.column_dimensions['E'].fill = fill_color
            ws.column_dimensions['K'].fill = fill_color
            ws.column_dimensions['H'].fill = fill_color

        try:
            wb.save(output_file)
            print(f"Файл сохраненq: {output_file}")
        except Exception as e:
            print(f"Ошибка при сохранении файла {output_file}: {e}")
            raise
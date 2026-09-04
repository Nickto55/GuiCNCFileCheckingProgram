import os
import openpyxl

from datetime import datetime

SEARCH_COLUMN_TITLE = "Содержимое"
PATH_COLUMN_TITLE = "Путь"
RESULT_COLUMN_TITLE = "Автор"
DATA_COLUMN_TITLE = "Дата последнего изменения"
KB_COLUMN_TITLE = "KБ"
DSE_NAME_COLUMN_TITLE = "Название ДСЕ"
FM_COLUMN_TITLE = "Fm"
FILES_NO_EXT_COLUMN_TITLE = "Файлы без расширения"
FILE_EXTENSIONS = [".nc", ".NC", ".h", ".H"]
SEARCH_FRAGMENT = "AVTOR"
IGNORED_SHEET = "ДСЕ по станкам"


def get_directory_size(directory_path: str) -> int:
    total_size = 0
    try:
        for dirpath, dirnames, filenames in os.walk(directory_path):
            for filename in filenames:
                filepath = os.path.join(dirpath, filename)
                try:
                    total_size += os.path.getsize(filepath)
                except OSError:
                    pass
    except OSError as e:
        print(f"Ошибка при получении размера папки: {e}")
        raise
    return total_size


def main(excel_file, db_handler=None, progress_tracker=None):
    try:
        wb = openpyxl.load_workbook(excel_file)
        sheets_to_process = [sheet for sheet in wb.sheetnames if sheet != IGNORED_SHEET]

        data_for_db = {}

        total_steps = 0
        if progress_tracker:
            for sheet_name in sheets_to_process:
                ws = wb[sheet_name]
                estimated_rows = ws.max_row - 1 if ws.max_row > 1 else 0
                total_steps += max(estimated_rows, 0)
            progress_tracker.set_total(total_steps)

        current_step = 0

        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]

            data_for_db[sheet_name] = []

            try:
                headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                headers = ()

            content_col, path_col, result_col, data_col, kb_col = None, None, None, None, None
            dse_col, fm_col, files_no_ext_col = None, None, None

            for idx, header in enumerate(headers):
                if header == SEARCH_COLUMN_TITLE: content_col = idx + 1
                elif header == PATH_COLUMN_TITLE: path_col = idx + 1
                elif header == DATA_COLUMN_TITLE: data_col = idx + 1
                elif header == KB_COLUMN_TITLE: kb_col = idx + 1
                elif header == RESULT_COLUMN_TITLE: result_col = idx + 1
                elif header == DSE_NAME_COLUMN_TITLE: dse_col = idx + 1
                elif header == FM_COLUMN_TITLE: fm_col = idx + 1
                elif header == FILES_NO_EXT_COLUMN_TITLE: files_no_ext_col = idx + 1

            if result_col is None:
                result_col = len(headers) + 1
                ws.cell(row=1, column=result_col, value=RESULT_COLUMN_TITLE)
            if data_col is None:
                data_col = len(headers) + 1
                ws.cell(row=1, column=data_col, value=DATA_COLUMN_TITLE)
            if kb_col is None:
                kb_col = len(headers) + 1
                ws.cell(row=1, column=kb_col, value=KB_COLUMN_TITLE)

            if not all([content_col, path_col]):
                if progress_tracker:
                    estimated_rows = ws.max_row - 1 if ws.max_row > 1 else 0
                    for _ in range(estimated_rows):
                        progress_tracker.update(current_step, f"Пропущен лист '{sheet_name}' (нет столбцов)")
                        current_step += 1
                continue

            rows_data = list(ws.iter_rows(min_row=2, values_only=False))
            for row_idx, row in enumerate(rows_data, start=2):
                file_name_cell = row[content_col - 1]
                path_cell = row[path_col - 1]
                progress_message = f"Обработка листа '{sheet_name}', строка {row_idx - 1}/{len(rows_data)}"

                if not (file_name_cell.value and path_cell.value):
                    if progress_tracker:
                        progress_tracker.update(current_step, progress_message + " - Пропущена (пустая)")
                        current_step += 1
                    if kb_col: row[kb_col - 1].value = ""
                    if data_col: row[data_col - 1].value = ""
                    if result_col: row[result_col - 1].value = ""
                    continue

                file_name = str(file_name_cell.value).strip()
                full_path = str(path_cell.value).strip()

                if not any(file_name.endswith(ext) for ext in FILE_EXTENSIONS):
                    if progress_tracker:
                        progress_tracker.update(current_step, progress_message + " - Пропущена (расширение)")
                        current_step += 1
                    if kb_col: row[kb_col - 1].value = ""
                    if data_col: row[data_col - 1].value = ""
                    continue

                if not os.path.exists(full_path):
                    if progress_tracker:
                        progress_tracker.update(current_step, progress_message + " - Ошибка (файл не найден)")
                        current_step += 1
                    if kb_col: row[kb_col - 1].value = "Файл не найден"
                    if data_col: row[data_col - 1].value = "Файл не найден"
                    if result_col: row[result_col - 1].value = "Файл не найден"
                    continue

                try:
                    fm_value = ""
                    if fm_col:
                        fm_cell_value = row[fm_col - 1].value
                        fm_value = str(fm_cell_value).strip() if fm_cell_value is not None else ""

                    files_no_ext_value = ""
                    if files_no_ext_col:
                        fne_cell_value = row[files_no_ext_col - 1].value
                        files_no_ext_value = str(fne_cell_value).strip() if fne_cell_value is not None else ""

                    if not files_no_ext_value:
                        files_no_ext_value = os.path.splitext(file_name)[0]
                        if files_no_ext_col:
                            row[files_no_ext_col - 1].value = files_no_ext_value

                    normalized_path = full_path.replace('\\', '/')
                    path_parts = normalized_path.split('/')

                    dse_name_value = ""
                    if dse_col and row[dse_col - 1].value:
                        dse_name_value = str(row[dse_col - 1].value).strip()
                    if not dse_name_value:
                        dse_name_value = path_parts[-2] if len(path_parts) >= 2 else file_name

                    if os.path.isfile(full_path):
                        size_bytes = os.path.getsize(full_path)
                    else:
                        size_bytes = get_directory_size(full_path)

                    kb_size = round(size_bytes / 1024)
                    if kb_col: row[kb_col - 1].value = kb_size

                    mod_date = datetime.fromtimestamp(os.path.getmtime(full_path)).strftime('%Y-%m-%d %H:%M:%S')
                    if data_col: row[data_col - 1].value = mod_date

                    avtor_value = "Не найден"
                    if os.path.isfile(full_path):
                        with open(full_path, 'r', encoding='utf-8', errors='ignore') as f:
                            for line in f:
                                if SEARCH_FRAGMENT in line:
                                    avtor_parts = line.strip().split("AVTOR:", 1)
                                    if len(avtor_parts) > 1:
                                        avtor_value = avtor_parts[1].strip().strip('()')
                                    break

                    if result_col: row[result_col - 1].value = avtor_value

                    # --- Накопление данных для filling_database ---
                    if sheet_name in normalized_path or sheet_name in full_path:
                        data_for_db[sheet_name].append({
                            "name": dse_name_value,               # из столбца "Название ДСЕ"
                            "content": avtor_value,               # из столбца "Автор"
                            "full_path": os.path.normpath(full_path),               # из столбца "Путь"
                            "fm_file": fm_value,                  # из столбца "Fm"
                            "files_without_extension": files_no_ext_value,  # из столбца "Файлы без расширения"
                            "last_modified_date": mod_date,       # из столбца "Дата последнего изменения"
                            "kb": kb_size                         # из столбца "KБ"
                        })

                    if progress_tracker:
                        progress_tracker.update(current_step, progress_message + " - Обработан")
                        current_step += 1

                except Exception as e:
                    if progress_tracker:
                        progress_tracker.update(current_step, progress_message + f" - Ошибка ({type(e).__name__})")
                        current_step += 1
                    if kb_col: row[kb_col - 1].value = f"Ошибка: {e}"
                    if data_col: row[data_col - 1].value = f"Ошибка: {e}"
                    if result_col: row[result_col - 1].value = f"Ошибка: {e}"

        wb.save(excel_file)
        print("Обработка завершена. Результаты записаны в исходный файл.")

        if db_handler is not None:
            try:
                db_handler.filling_database(data_for_db)
                print("База данных успешно заполнена.")
            except Exception as db_error:
                print(f"author_verification_program.Ошибка при заполнении базы данных: {db_error}")
                raise
        else:
            print("db_handler не передан, запись в базу данных пропущена.")

        if progress_tracker and current_step < total_steps:
            for i in range(current_step, total_steps):
                progress_tracker.update(i, "Завершение...")
            if total_steps > 0:
                progress_tracker.update(total_steps - 1, "Обработка авторов завершена")

    except Exception as e:
        print(f"Критическая ошибка в main AuthorVerificationProgram: {e}")
        raise
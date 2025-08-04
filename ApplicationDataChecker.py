import os
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn, Filters


def auto_fit_columns(sheet):
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        adjusted_width = (max_length + 2)  # Немного увеличим для красоты
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
    # sheet.column_dimensions['A'].width = 5 # Убираем, так как задается позже


def main(fileName: str, progress_tracker=None):
    """
    Создает сводную таблицу "ДСЕ по станкам" в Excel-файле.

    Args:
        fileName (str): Имя Excel-файла (без пути).
        progress_tracker (ProgressTracker, optional): Объект для отслеживания прогресса.

    Returns:
        str: Полный путь к обработанному файлу.
    """
    try:
        # Получаем текущую директорию
        base_dir = os.getcwd()
        file_path = os.path.join(base_dir, f'{fileName}')

        # --- Логика подсчета общего количества шагов для прогресс-бара ---
        total_steps = 0
        if progress_tracker:
            try:
                # Предварительно открыть книгу, чтобы оценить количество листов и данных
                # Это нужно для точного подсчета шагов.
                temp_wb = openpyxl.load_workbook(file_path, read_only=True)  # Открываем в режиме чтения для оценки
                temp_sheets = [sheet for sheet in temp_wb.sheetnames if sheet != "ДСЕ по станкам"]

                # Примерная оценка: количество листов (для сбора данных) +
                # количество уникальных файлов (для заполнения).
                # Для упрощения, можно оценить как сумму max_row по всем листам.
                estimated_data_collection_steps = len(temp_sheets)
                estimated_fill_steps = 0
                unique_files_estimate = set()
                for sheet_name in temp_sheets:
                    temp_ws = temp_wb[sheet_name]
                    # Оцениваем количество строк с данными в колонке B (Файл)
                    # Используем iter_rows для большей эффективности при оценке
                    for row in temp_ws.iter_rows(min_row=2, max_col=2, values_only=True):
                        if row and row[0]:  # row[0] это значение ячейки B
                            unique_files_estimate.add(str(row[0]).strip())

                estimated_fill_steps = len(unique_files_estimate)
                total_steps = estimated_data_collection_steps + estimated_fill_steps
                temp_wb.close()  # Закрываем временную книгу

            except Exception as e:
                print(f"Предупреждение: Не удалось оценить количество шагов для прогресса: {e}")
                # Если не удалось оценить, используем фиктивное значение или 0
                total_steps = 0  # Или какое-то приблизительное значение, если можно

            progress_tracker.set_total(total_steps)
        # --- Конец логики подсчета шагов ---

        current_step = 0

        # Открываем книгу для записи
        wb = openpyxl.load_workbook(file_path)

        # Удаляем старый лист "ДСЕ по станкам", если есть
        if "ДСЕ по станкам" in wb.sheetnames:
            del wb["ДСЕ по станкам"]
            if progress_tracker:
                current_step += 1
                progress_tracker.update(current_step, "Удален старый лист 'ДСЕ по станкам'")

        ws_summary = wb.create_sheet(title="ДСЕ по станкам")
        if progress_tracker:
            current_step += 1
            progress_tracker.update(current_step, "Создан лист 'ДСЕ по станкам'")

        # Словарь: ключ - имя файла, значение - словарь {станок: ссылка}
        file_data = {}

        # Перечисляем все листы, кроме "ДСЕ по станкам"
        sheets = [sheet for sheet in wb.sheetnames if sheet != "ДСЕ по станкам"]
        # column_indices = [i for i in range(1, 2+len(sheets))] # Не используется напрямую для фильтра

        # Собираем данные из всех листов
        for sheet_name in sheets:
            ws = wb[sheet_name]
            # print(f"Сбор данных с листа: {sheet_name}") # Для дебага
            for row in range(2, ws.max_row + 1):  # Начиная со второй строки
                file_name_cell_value = ws.cell(row=row, column=2).value  # Значение ячейки B
                link_cell_value = ws.cell(row=row, column=4).value  # Значение ячейки D

                if file_name_cell_value and link_cell_value:
                    file_name_str = str(file_name_cell_value).strip()
                    link_str = str(link_cell_value).strip()
                    if file_name_str not in file_data:
                        file_data[file_name_str] = {}
                    file_data[file_name_str][sheet_name] = link_str  # Сохраняем ссылку как строку

            # Обновляем прогресс после обработки листа
            if progress_tracker:
                current_step += 1
                progress_tracker.update(current_step, f"Собраны данные с листа '{sheet_name}'")

        # Формируем заголовки
        ws_summary["B1"] = "Файл"
        if progress_tracker:
            current_step += 1
            progress_tracker.update(current_step, "Сформированы заголовки")

        col_offset = 3  # Начинаем с C (колонка 3)
        for idx, sheet_name in enumerate(sheets):
            col_letter = get_column_letter(col_offset + idx)
            ws_summary[f"{col_letter}1"] = sheet_name

        # Заполняем таблицу
        row_idx = 2
        all_files = sorted(file_data.keys())

        # Обновляем диапазон фильтра динамически
        if all_files:
            last_col_letter = get_column_letter(col_offset + len(sheets) - 1)
            last_row = len(all_files) + 1  # +1 for header
            data_range = f"B1:{last_col_letter}{last_row}"
        else:
            # Если файлов нет, создаем минимальный диапазон
            last_col_letter = get_column_letter(col_offset + max(len(sheets) - 1, 0))
            data_range = f"B1:{last_col_letter}1"  # Только заголовок

        for file_name in all_files:
            ws_summary[f"B{row_idx}"] = file_name

            for idx, sheet_name in enumerate(sheets):
                col_letter = get_column_letter(col_offset + idx)
                cell_ref = f"{col_letter}{row_idx}"
                cell_obj = ws_summary[cell_ref]  # Получаем объект ячейки для работы

                if sheet_name in file_data.get(file_name, {}):
                    full_file_path = file_data[file_name][sheet_name]  # Это путь к .nc/.h файлу
                    # Гиперссылка должна вести к самому файлу .nc/.h
                    cell_obj.hyperlink = full_file_path

                    # Получаем дату изменения файла .nc/.h
                    try:
                        modification_time = os.path.getmtime(full_file_path)
                        file_date = datetime.fromtimestamp(modification_time)
                        cell_obj.value = file_date
                        cell_obj.number_format = 'dd.mm.yyyy'
                    except OSError as e:
                        # Если файл не найден или ошибка доступа
                        print(f"Предупреждение: Не удалось получить дату для файла {full_file_path}: {e}")
                        cell_obj.value = "Ошибка даты"
                        cell_obj.number_format = 'General'  # Сбрасываем формат

                    except Exception as e:
                        print(f"Ошибка при обработке даты файла {full_file_path}: {e}")
                        cell_obj.value = "Ошибка"
                        cell_obj.number_format = 'General'

                    # Стили для существующего файла
                    cell_obj.font = Font(color="0000FF", underline="single")  # Добавил underline для гиперссылок
                    cell_obj.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    # Исправлена логика границ
                    thin_side = Side(border_style="thin", color="000000")  # Используем более стандартные границы
                    cell_obj.border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
                else:
                    cell_obj.value = "-"  # Используем простой дефис
                    cell_obj.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    # Сброс стилей для отсутствующего файла
                    cell_obj.font = Font()  # Сброс шрифта
                    cell_obj.border = Border()  # Сброс границ

            row_idx += 1

            # Обновляем прогресс после обработки строки (файла)
            if progress_tracker:
                current_step += 1
                progress_tracker.update(current_step, f"Заполнена строка для файла '{file_name}'")

        # создание фильтра и его настройка
        ws_summary.auto_filter.ref = data_range
        # Создаем список индексов колонок для фильтрации (B=2, C=3, ...)
        column_indices_for_filter = list(range(2, col_offset + len(sheets)))  # 2 (B) до последней колонки с данными
        for col_index in column_indices_for_filter:
            filter_column = FilterColumn(colId=col_index - 1)  # индексы в OpenPyXL начинаются с 0
            filters = Filters()
            filter_column.filters = filters
            ws_summary.auto_filter.filterColumn.append(filter_column)

        if progress_tracker:
            current_step += 1
            progress_tracker.update(current_step, "Настроен автофильтр")

        auto_fit_columns(ws_summary)
        ws_summary.column_dimensions['B'].width = 20
        # Закрепление первой строки
        ws_summary.freeze_panes = 'B2'  # Исправлено: закрепляем от B2, чтобы колонка A оставалась

        if progress_tracker:
            current_step += 1
            progress_tracker.update(current_step, "Настроены ширины столбцов и закрепление")

        def saveTry(wb, file_path, current_step_tracker, current_step_count):
            try:
                wb.save(file_path)
                if current_step_tracker:
                    current_step_count += 1
                    current_step_tracker.update(current_step_count, "Файл успешно сохранен")
            except PermissionError as e:
                print(f"Ошибка доступа при сохранении файла: {e}")
                # В GUI лучше обрабатывать это через callback или исключение
                # Пока просто пробуем снова
                print("Возможно, файл открыт в другом приложении. Закройте его и нажмите Enter...")
                input("Нажмите Enter для повторной попытки сохранения...")
                saveTry(wb, file_path, current_step_tracker, current_step_count)
            except Exception as e:
                print(f"Ошибка при сохранении файла: {e}")
                raise  # Перебрасываем, если это не PermissionError

        saveTry(wb, file_path, progress_tracker, current_step)

        if progress_tracker and current_step < total_steps:
            # Обновляем оставшиеся шаги, если они есть (например, если оценка была неточной)
            for i in range(current_step, total_steps):
                progress_tracker.update(i, "Завершение...")
            # progress_tracker.update(total_steps - 1, "Создание сводной таблицы завершено")

        print(f"Сводная таблица создана и сохранена в {file_path}")
        return file_path

    except Exception as e:
        print(f"Критическая ошибка в main ApplicationDataChecker: {e}")
        raise  # Перебрасываем исключение, чтобы GUI мог его поймать

# if __name__ == '__main__':
#     # Пример вызова без прогресс-трекера
#     # main("ваш_файл.xlsx")
#     pass
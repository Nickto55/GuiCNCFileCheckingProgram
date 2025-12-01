import os
from datetime import datetime

import openpyxl

# Параметры (можно оставить как константы или передавать как аргументы, если нужно)
SEARCH_COLUMN_TITLE = "Содержимое"  # Заголовок столбца с именами файлов
PATH_COLUMN_TITLE = "Путь"  # Заголовок столбца с полными путями
RESULT_COLUMN_TITLE = "Автор"  # Заголовок для нового столбца с результатами
DATA_COLUMN_TITLE = "Дата последнего изменения"  # Заголовок для нового столбца с результатами
KB_COLUMN_TITLE = "KБ"  # Заголовок для нового столбца с результатами вес
FILE_EXTENSIONS = [".nc", ".NC", ".h", ".H"]  # Поддерживаемые расширения
SEARCH_FRAGMENT = "AVTOR"  # Фрагмент строки для поиска
IGNORED_SHEET = "ДСЕ по станкам"  # Лист, который нужно пропустить


def get_directory_size(directory_path: str) -> int:
    """
    Возвращает суммарный размер всех файлов в папке (включая вложенные папки) в байтах.

    Args:
        directory_path (str): Путь к папке.

    Returns:
        int: Размер папки в байтах.
    """
    total_size = 0
    try:
        for dirpath, dirnames, filenames in os.walk(directory_path):
            for filename in filenames:
                filepath = os.path.join(dirpath, filename)
                try:
                    total_size += os.path.getsize(filepath)
                except OSError:
                    # Игнорируем файлы, к которым нет доступа
                    pass
    except OSError as e:
        print(f"Ошибка при получении размера папки: {e}")
        raise
    return total_size


def main(excel_file, progress_tracker=None):
    """
    Основная функция для поиска авторов в файлах и записи их в Excel.

    Args:
        excel_file (str): Путь к Excel-файлу для обработки.
        progress_tracker (ProgressTracker, optional): Объект для отслеживания прогресса.
    """
    try:
        # Открываем Excel-файл
        wb = openpyxl.load_workbook(excel_file)
        sheets_to_process = [sheet for sheet in wb.sheetnames if sheet != IGNORED_SHEET]

        # --- Логика подсчета общего количества шагов для прогресс-бара ---
        total_steps = 0
        if progress_tracker:
            for sheet_name in sheets_to_process:
                ws = wb[sheet_name]
                # Примерно считаем количество строк данных (вычитаем 1 для заголовка)
                estimated_rows = ws.max_row - 1 if ws.max_row > 1 else 0
                total_steps += max(estimated_rows, 0)  # Убедиться, что не отрицательное

            progress_tracker.set_total(total_steps)
        # --- Конец логики подсчета шагов ---

        current_step = 0

        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            # print(f"Обработка листа: {sheet_name}") # Можно убрать или оставить для дебага

            # Находим номера столбцов по заголовкам
            headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))  # Получаем первую строку (заголовки)
            content_col, path_col, result_col, data_col, kb_col = None, None, None, None, None

            for idx, header in enumerate(headers):
                if header == SEARCH_COLUMN_TITLE:
                    content_col = idx + 1  # Индексы в OpenPyXL начинаются с 1
                elif header == PATH_COLUMN_TITLE:
                    path_col = idx + 1
                elif header == DATA_COLUMN_TITLE:
                    data_col = idx + 1
                elif header == KB_COLUMN_TITLE:
                    kb_col = idx + 1
                elif header == RESULT_COLUMN_TITLE:
                    result_col = idx + 1  # Если столбец уже существует

            # Если заголовок результата не найден — добавляем его
            if result_col is None:
                result_col = len(headers) + 1
                ws.cell(row=1, column=result_col, value=RESULT_COLUMN_TITLE)

            # Если заголовок даты не найден — добавляем его
            if data_col is None:
                data_col = len(headers) + 1
                ws.cell(row=1, column=data_col, value=DATA_COLUMN_TITLE)

            # Если заголовок размера в КБ не найден — добавляем его
            if kb_col is None:
                kb_col = len(headers) + 1
                ws.cell(row=1, column=kb_col, value=KB_COLUMN_TITLE)

            # Проверяем, что найдены все нужные столбцы
            if not all([content_col, path_col]):
                print(f"Столбцы '{SEARCH_COLUMN_TITLE}' или '{PATH_COLUMN_TITLE}' не найдены на листе '{sheet_name}'")
                # Даже если столбцы не найдены, увеличиваем счётчик шагов на оценочное количество строк
                if progress_tracker:
                    estimated_rows = ws.max_row - 1 if ws.max_row > 1 else 0
                    for _ in range(estimated_rows):
                        progress_tracker.update(current_step, f"Пропущен лист '{sheet_name}' (нет столбцов)")
                        current_step += 1
                continue  # Переходим к следующему листу

            # Обрабатываем каждую строку
            # Используем значения, чтобы получить реальное количество строк с данными
            rows_data = list(ws.iter_rows(min_row=2, values_only=False))  # Загружаем данные
            for row_idx, row in enumerate(rows_data, start=2):  # Начинаем со второй строки
                file_name_cell = row[content_col - 1]
                path_cell = row[path_col - 1]

                # Сообщение для прогресса
                progress_message = f"Обработка листа '{sheet_name}', строка {row_idx - 1}/{len(rows_data)}"

                if not (file_name_cell.value and path_cell.value):
                    # Пропускаем пустые ячейки
                    if progress_tracker:
                        progress_tracker.update(current_step, progress_message + " - Пропущена (пустая)")
                        current_step += 1
                    # Устанавливаем значения по умолчанию для пустых строк
                    if kb_col:
                        kb_cell = row[kb_col - 1]
                        kb_cell.value = ""
                    if data_col:
                        data_cell = row[data_col - 1]
                        data_cell.value = ""
                    if result_col:
                        result_cell = row[result_col - 1]
                        result_cell.value = ""
                    continue

                file_name = str(file_name_cell.value).strip()
                full_path = str(path_cell.value).strip()

                # Проверяем расширение файла
                if not any(file_name.endswith(ext) for ext in FILE_EXTENSIONS):
                    # Пропускаем файлы с неподдерживаемыми расширениями
                    if progress_tracker:
                        progress_tracker.update(current_step, progress_message + " - Пропущена (расширение)")
                        current_step += 1
                    # Устанавливаем значения по умолчанию для строк с неподдерживаемым расширением
                    if kb_col:
                        kb_cell = row[kb_col - 1]
                        kb_cell.value = ""
                    if data_col:
                        data_cell = row[data_col - 1]
                        data_cell.value = ""
                    continue

                # Проверяем существование файла
                if not os.path.exists(full_path):
                    # print(f"Файл не существует: {full_path}")
                    if progress_tracker:
                        progress_tracker.update(current_step, progress_message + " - Ошибка (файл не найден)")
                        current_step += 1
                    # Устанавливаем значения по умолчанию для несуществующих файлов
                    if kb_col:
                        kb_cell = row[kb_col - 1]
                        kb_cell.value = "Файл не найден"
                    if data_col:
                        data_cell = row[data_col - 1]
                        data_cell.value = "Файл не найден"
                    if result_col:
                        result_cell = row[result_col - 1]
                        result_cell.value = "Файл не найден"
                    continue

                try:
                    # Получаем размер файла или директории
                    if os.path.isfile(full_path):
                        size_bytes = os.path.getsize(full_path)
                    else:
                        size_bytes = get_directory_size(full_path)
                    size_kb = round(size_bytes / 1024)  # Размер в КБ, округленный до целого

                    # Записываем размер в КБ в соответствующую ячейку
                    kb_cell = row[kb_col - 1]
                    kb_cell.value = size_kb

                    # Получаем дату последнего изменения
                    mod_time = os.path.getmtime(full_path)
                    mod_date = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d %H:%M:%S')

                    # Записываем дату в соответствующую ячейку
                    data_cell = row[data_col - 1]
                    data_cell.value = mod_date

                    with open(full_path, 'r', encoding='utf-8', errors='ignore') as f:  # Добавлен encoding и errors
                        author_found = False
                        for line in f:
                            if SEARCH_FRAGMENT in line:
                                avtor_line = line.strip()
                                # Извлекаем значение после "AVTOR:"
                                # Пример: AVTOR:(Иванов И.И.)
                                avtor_value = avtor_line.split("AVTOR:", 1)
                                if len(avtor_value) > 1:
                                    avtor_value = avtor_value[1].strip()
                                    # Убираем скобки, если они есть в начале и конце
                                    if avtor_value.startswith('(') and avtor_value.endswith(')'):
                                        avtor_value = avtor_value[1:-1]

                                    # Записываем результат в столбец "Автор"
                                    result_cell = row[result_col - 1]
                                    result_cell.value = avtor_value
                                    author_found = True
                                    break  # Найден автор, выходим из цикла чтения файла
                        if not author_found:
                            result_cell = row[result_col - 1]
                            result_cell.value = "Не найден"
                    # Обновляем прогресс после обработки файла
                    if progress_tracker:
                        progress_tracker.update(current_step, progress_message + " - Обработан")
                        current_step += 1

                except Exception as e:
                    # print(f"Ошибка при открытии файла {full_path}: {e}")
                    if progress_tracker:
                        progress_tracker.update(current_step, progress_message + f" - Ошибка ({type(e).__name__})")
                        current_step += 1
                    # Записываем ошибку в соответствующие ячейки
                    kb_cell = row[kb_col - 1]
                    kb_cell.value = f"Ошибка: {e}"
                    data_cell = row[data_col - 1]
                    data_cell.value = f"Ошибка: {e}"
                    result_cell = row[result_col - 1]
                    result_cell.value = f"Ошибка: {e}"

        # Сохраняем изменения
        wb.save(excel_file)
        print("Обработка завершена. Результаты записаны в исходный файл.")
        # Если прогресс дошел не до конца (например, меньше строк, чем ожидалось), обновим до 100%
        if progress_tracker and current_step < total_steps:
            # Обновляем оставшиеся шаги
            for i in range(current_step, total_steps):
                progress_tracker.update(i, "Завершение...")
            progress_tracker.update(total_steps - 1,
                                    "Обработка авторов завершена")  # Убедимся, что последний шаг отмечен

    except Exception as e:
        print(f"Критическая ошибка в main AuthorVerificationProgram: {e}")
        raise  # Перебрасываем исключение, чтобы GUI мог его поймать
